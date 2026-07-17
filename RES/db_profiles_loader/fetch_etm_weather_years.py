"""Fetch ETM "weather year" curves via pyetm and append them to db_profiles.xlsx.

Implements the task in etm_weather_years_pyetm_instructions.md: for each of the
1987 / 1997 / 2004 ETM weather-year settings (plus "default" as a sanity check),
build hourly + quarter-hourly profile sheets matching the column layout of the
existing profiles_2025 sheet, and append them as new sheets to a copy of
db_profiles.xlsx.

Discovered facts this script relies on (see NOTES.md for the full trail):
- The scenario input key for weather year is "settings_weather_curve_set", an
  EnumInput with permitted_values ["default", "1987", "1997", "2004"], not
  disabled for area_code "nl2023". Verified programmatically, not guessed.
- session.get_hourly_curve(name) only returns hourly (8760-row) curves - there
  is no resolution parameter. So the quarter-hourly sheets in this script are
  produced by interpolation, unlike the original profiles_2025 sheet, where a
  direct comparison against profiles_2025_h showed wind (a production column)
  is independently resampled at 15-minute resolution rather than interpolated
  from the hourly column. This is a real, documented limitation of the ETM
  pull, not an oversight.
- Only 10 hourly-curve identifiers are attached for this scenario
  (session.hourly_output_curves.attached_keys()): electricity_profiles,
  district_heating_profiles, hydrogen_profiles, network_gas_profiles,
  electricity_price, agriculture_heat, household_heat, buildings_heat,
  residual_load, hydrogen_integral_cost. There is no curve literally named
  "industry_heat" - but industry heat demand IS present, as final energy
  demand for network gas / hydrogen within network_gas_profiles and
  hydrogen_profiles, split per subsector (see INDUSTRY_*_H_*_COLS below).
  Chemicals subsectors only expose a "_non_energetic" (feedstock, not heat)
  final-demand node for gas/hydrogen - their actual heat comes through a
  separate "_burner_" supply node instead. Getting this wrong the first time
  (assuming "no industry_heat curve" meant "no data") is why these two
  columns are double-checked explicitly in NOTES.md.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime, timezone

import numpy as np
import openpyxl
import pandas as pd
import pyetm
from pyetm import Session
from pyetm.config import reload_configuration

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger("fetch_etm_weather_years")

AREA_CODE = "nl2023"
END_YEAR = 2050
WEATHER_INPUT_KEY = "settings_weather_curve_set"
WEATHER_YEARS = ["default", "1987", "1997", "2004"]

SOURCE_XLSX = "../db_profiles.xlsx"
OUTPUT_XLSX = "db_profiles_with_weather_years.xlsx"

# Column layout of profiles_2025 / profiles_2024 / profiles_2024_366_days (the
# "current" layout per the instructions §0). profiles_2023* additionally has
# wind_e_prod_normalized_herwijnen / _ell, which are intentionally NOT
# reproduced here - out of scope, see NOTES.md.
COLUMN_ORDER = [
    "t_h",
    "wind_e_prod_normalized",
    "solar_e_prod_south35deg_normalized",
    "solar_e_prod_eastwest15deg_normalized",
    "ambientTemperature_degC",
    "day_ahead_price_eur_p_mwh",
    "house_e_demand_other",
    "house_h_demand_hot_water",
    "building_e_demand_other",
    "building_h_demand",
    "house_cooking_demand",
    "industry_steel_e_demand",
    "industry_steel_h_demand",
    "industry_other_e_demand",
    "industry_other_h_demand",
    "logistics_fleet_e_hgv",
    "wind_e_prod_normalized_hvh",
    "wind_e_prod_normalized_heibloem",
    "co2_factor_kg_per_kwh",
]

# "max" -> production-style: divide by the series max (peak hour = 1)
# "sum" -> demand-style: divide by the annual sum (all hours add up to 1)
# "none" -> leave as absolute value / not sourced from ETM
NORMALIZATION = {
    "wind_e_prod_normalized": "max",
    "solar_e_prod_south35deg_normalized": "max",
    "solar_e_prod_eastwest15deg_normalized": "max",
    "ambientTemperature_degC": "none",
    "day_ahead_price_eur_p_mwh": "none",
    "house_e_demand_other": "sum",
    "house_h_demand_hot_water": "sum",
    "building_e_demand_other": "sum",
    "building_h_demand": "sum",
    "house_cooking_demand": "sum",
    "industry_steel_e_demand": "sum",
    "industry_steel_h_demand": "sum",
    "industry_other_e_demand": "sum",
    "industry_other_h_demand": "sum",
    "logistics_fleet_e_hgv": "sum",
    "wind_e_prod_normalized_hvh": "max",
    "wind_e_prod_normalized_heibloem": "max",
    "co2_factor_kg_per_kwh": "none",
}

WIND_INLAND_COL = "energy_power_wind_turbine_inland.output (MW)"
# Genuinely distinct from WIND_INLAND_COL (different ETM production node) - added on
# request as its own column, alongside wind_e_prod_normalized (which stays sourced from
# WIND_INLAND_COL, duplicated across _hvh/_heibloem as before, since ETM has no
# per-location split for those original db_profiles.xlsx columns). Not to be confused
# with "energy_power_hybrid_wind_turbine_offshore.output (MW)", a separate hybrid node.
WIND_OFFSHORE_COL = "energy_power_wind_turbine_offshore.output (MW)"

SOLAR_COLS = [
    "energy_power_solar_pv_solar_radiation.output (MW)",
    "households_solar_pv_solar_radiation.output (MW)",
    "households_solar_pvt_solar_radiation.output (MW)",
    "buildings_solar_pv_solar_radiation.output (MW)",
]

# house_e_demand_other / building_e_demand_other = final demand for appliances + lighting
# only, on request - deliberately excludes cooling (households_cooling_*/
# buildings_cooling_*.input (MW)). Cooling used to be bundled in here, which is exactly
# why these two columns showed real weather-year variation despite being expected as a
# "fixed profile" - cooling is temperature-driven (verified: ~3x more cooling in a hot
# year vs. a cold one), appliances/lighting are not (verified: byte-for-byte identical
# across weather years). Excluding cooling makes both columns genuinely
# weather-independent, matching what they're meant to represent. All columns confirmed
# in "(MW)" throughout electricity_profiles - no MW/MWh mix found when checked directly
# (electricity_profiles/household_heat/buildings_heat/network_gas_profiles/
# hydrogen_profiles all use "(MW)" exclusively; the one non-MW column across all of
# them is "deficit", unused here).
HOUSE_E_OTHER_COLS = [
    "households_final_demand_for_appliances_electricity.input (MW)",
    "households_final_demand_for_lighting_electricity.input (MW)",
]

BUILDING_E_OTHER_COLS = [
    "buildings_appliances_electricity.input (MW)",
    "buildings_lighting_efficient_fluorescent_electricity.input (MW)",
    "buildings_lighting_led_electricity.input (MW)",
    "buildings_lighting_standard_fluorescent_electricity.input (MW)",
]

HOUSE_COOKING_COL = "households_final_demand_for_cooking_electricity.input (MW)"

LOGISTICS_HGV_COL = "transport_truck_using_electricity.input (MW)"

HOUSE_HOT_WATER_COL = "households_useful_demand_for_hot_water_after_solar_heater.input.demand (MW)"

BUILDING_HEAT_COLS = [
    "buildings_useful_demand_for_space_heating_buildings_future_after_solar_thermal.input.demand (MW)",
    "buildings_useful_demand_for_space_heating_buildings_present_after_solar_thermal.input.demand (MW)",
]

# household_heat's per-(house type x construction era) space heating useful-demand
# columns - the household equivalent of BUILDING_HEAT_COLS above. Listed explicitly
# (not just pattern-matched at call time) so this is visible/greppable next to
# HOUSE_HOT_WATER_COL/BUILDING_HEAT_COLS rather than buried inside fetch_weather_year().
# fetch_weather_year() still cross-checks this list against household_heat.columns at
# runtime and warns if they diverge (e.g. a future ETM dataset changing the house-type/
# era breakdown), rather than silently using whatever pattern-match happens to find.
HOUSE_SPACE_HEATING_COLS = [
    f"households_useful_demand_for_space_heating_{house_type}_{era}.input.demand (MW)"
    for house_type in ["apartments", "terraced_houses", "semi_detached_houses", "detached_houses"]
    for era in ["future", "2005_present", "1985_2004", "1965_1984", "1945_1964", "before_1945"]
]

# Industry heat is NOT exposed as a dedicated "industry_heat" hourly curve (unlike
# household_heat/buildings_heat) - it shows up as final energy demand for network gas
# and hydrogen within network_gas_profiles / hydrogen_profiles, split per subsector.
# Two node types matter here, and must not be confused:
#   - "industry_final_demand_for_metal_{steel,aluminium,other_metals}_{network_gas,hydrogen}.input"
#     -> real energetic (heat) demand for metals subsectors, no non_energetic variant exists.
#   - "industry_chemicals_{fertilizers,other,refineries}_burner_{network_gas,hydrogen}.input"
#     -> the heat-producing burner for chemicals; the "final_demand_for_chemical_*" nodes for
#     these three subsectors are EXCLUSIVELY "_non_energetic" (feedstock, e.g. ammonia/plastics
#     feedstock), which is NOT heat and must be excluded.
#   - "industry_other_{food,paper}_burner_{network_gas,hydrogen}.input" -> same burner pattern.
#   - "industry_final_demand_for_other_non_specified_{network_gas,hydrogen}.input" -> the plain
#     energetic variant (there's also a "_non_energetic" sibling, excluded).
# CHP inputs (industry_chp_*) and "other_final_demand_network_gas" (not industry-prefixed) are
# deliberately excluded - CHP is a shared cross-sector cogen node, not attributable to a single
# subsector, and "other_final_demand_*" isn't an industry node at all.
INDUSTRY_STEEL_H_GAS_COLS = ["industry_final_demand_for_metal_steel_network_gas.input (MW)"]
INDUSTRY_STEEL_H_H2_COLS = ["industry_final_demand_for_metal_steel_hydrogen.input (MW)"]

# Per-subsector breakdown of "industry_other_h_demand" - added on request, as EXTRA columns
# appended after the standard COLUMN_ORDER layout (see WEATHER_SHEET_COLUMNS below). Each
# entry is (output_column_name, gas_cols, h2_cols). industry_other_h_demand itself is no
# longer a separately-hardcoded column list - it's computed as the sum of these subsector
# series (see fetch_weather_year), so the two can never drift apart.
INDUSTRY_OTHER_H_SUBSECTORS: list[tuple[str, list[str], list[str]]] = [
    (
        "industry_chemicals_fertilizers_h_demand",
        ["industry_chemicals_fertilizers_burner_network_gas.input (MW)"],
        ["industry_chemicals_fertilizers_burner_hydrogen.input (MW)"],
    ),
    (
        "industry_chemicals_other_h_demand",
        ["industry_chemicals_other_burner_network_gas.input (MW)"],
        ["industry_chemicals_other_burner_hydrogen.input (MW)"],
    ),
    (
        "industry_chemicals_refineries_h_demand",
        ["industry_chemicals_refineries_burner_network_gas.input (MW)"],
        ["industry_chemicals_refineries_burner_hydrogen.input (MW)"],
    ),
    (
        "industry_metal_aluminium_h_demand",
        ["industry_final_demand_for_metal_aluminium_network_gas.input (MW)"],
        ["industry_final_demand_for_metal_aluminium_hydrogen.input (MW)"],
    ),
    (
        "industry_metal_other_metals_h_demand",
        ["industry_final_demand_for_metal_other_metals_network_gas.input (MW)"],
        ["industry_final_demand_for_metal_other_metals_hydrogen.input (MW)"],
    ),
    (
        "industry_other_non_specified_h_demand",
        ["industry_final_demand_for_other_non_specified_network_gas.input (MW)"],
        ["industry_final_demand_for_other_non_specified_hydrogen.input (MW)"],
    ),
    (
        "industry_food_h_demand",
        ["industry_other_food_burner_network_gas.input (MW)"],
        ["industry_other_food_burner_hydrogen.input (MW)"],
    ),
    (
        "industry_paper_h_demand",
        ["industry_other_paper_burner_network_gas.input (MW)"],
        ["industry_other_paper_burner_hydrogen.input (MW)"],
    ),
]
EXTRA_SUBSECTOR_COLUMNS = [name for name, _, _ in INDUSTRY_OTHER_H_SUBSECTORS]
for _col in EXTRA_SUBSECTOR_COLUMNS:
    NORMALIZATION[_col] = "sum"

# Household space heating demand - added on request. The original db_profiles.xlsx
# layout covers household hot water (house_h_demand_hot_water) and building space
# heating (building_h_demand), but never household space heating - a real gap, since
# space heating is normally the largest single household heat end-use. Sourced from
# household_heat via HOUSE_SPACE_HEATING_COLS above (24 columns: apartments/
# terraced/semi_detached/detached x future/2005_present/1985_2004/1965_1984/
# 1945_1964/before_1945). Same ".input.demand" convention as
# house_h_demand_hot_water/building_h_demand - pre-technology-split useful heat demand,
# not any one heating technology's output.
#
# wind_e_prod_normalized_offshore - added on request, alongside wind_e_prod_normalized
# (inland). A genuinely separate ETM production node (WIND_OFFSHORE_COL above), not a
# duplicate like _hvh/_heibloem are of the inland curve.
EXTRA_HOUSEHOLD_COLUMNS = ["house_h_demand_space_heating"]
NORMALIZATION["house_h_demand_space_heating"] = "sum"

EXTRA_WIND_COLUMNS = ["wind_e_prod_normalized_offshore"]
NORMALIZATION["wind_e_prod_normalized_offshore"] = "max"

# NBNL 2025 scenarios (Netbeheer Nederland's four published narratives, see
# https://energytransitionmodel.com/ "Featured scenarios") x 4 target years. Provides
# ETM's own market-clearing price (electricity_price curve) per scenario/year as a
# proxy for future EPEX day-ahead price, since real EPEX data only exists historically
# (2023-2025, see day_ahead_price_eur_p_mwh) and these are genuine published scenarios
# rather than an arbitrary "make up a future price" exercise.
#
# IMPORTANT: these scenarios live on ETM's version-pinned "2025-01" engine
# (https://2025-01.engine.energytransitionmodel.com), not the default "pro" engine used
# for the rest of this script - the scenario IDs below don't exist on "pro" (404).
# Resolved by following each saved_scenario's redirect chain from the ETM homepage
# (energytransitionmodel.com -> my.energytransitionmodel.com/saved_scenarios/<id> ->
# a "load" link containing "scenario_id=<engine id>&title=NBNL_<code>_<year>") - not
# guessed, and not otherwise exposed by pyetm itself.
#
# Scenario codes match ETM's own scenario titles (NBNL_KM_2030 etc.):
#   km = Koersvaste Middenweg, ev = Eigen Vermogen, gb = Gezamenlijke Balans,
#   ha = Horizon Aanvoer.
NBNL_ENGINE_ENVIRONMENT = "2025-01"
NBNL_SCENARIOS: dict[str, dict[str, object]] = {
    "km": {"name": "Koersvaste Middenweg", "years": {2030: 12102, 2035: 501, 2040: 504, 2050: 506}},
    "ev": {"name": "Eigen Vermogen", "years": {2030: 509, 2035: 511, 2040: 513, 2050: 515}},
    "gb": {"name": "Gezamenlijke Balans", "years": {2030: 517, 2035: 519, 2040: 521, 2050: 523}},
    "ha": {"name": "Horizon Aanvoer", "years": {2030: 525, 2035: 527, 2040: 529, 2050: 531}},
}
NBNL_TARGET_YEARS = [2030, 2035, 2040, 2050]

# Every NBNL scenario ships with its own custom curve uploads under a "weather/"
# namespace (verified identical across all 16 scenario/year combinations checked:
# km/ha/ev/gb at multiple years). These are literal pinned curves that OVERRIDE the
# parametric settings_weather_curve_set input entirely - confirmed by comparing wind
# production hour-by-hour across weather years with these curves still attached: the
# curve was byte-for-byte IDENTICAL regardless of settings_weather_curve_set (a real
# bug caught only because the resulting price barely moved between weather years, far
# less than the "Dunkelflaute" framing implies - a user's suspicion, not something we'd
# have caught from the numbers alone). After removing these 8 curves, wind production
# shape correlation between weather years drops from 1.0 (identical) to 0.03-0.19 (real
# reshuffling), and default-vs-1987 mean price swings from a suspicious ~1% to a
# plausible ~47%. The insulation_* curves under the same "weather/" namespace are left
# alone - those are a technology/retrofit-rate assumption, not tied to which historical
# year's weather is selected.
NBNL_WEATHER_CUSTOM_CURVES = [
    "weather/wind_offshore_baseline",
    "weather/wind_coastal_baseline",
    "weather/wind_inland_baseline",
    "weather/solar_pv_profile_1",
    "weather/solar_thermal",
    "weather/air_temperature",
    "weather/buildings_heating",
    "weather/agriculture_heating",
]


def nbnl_price_column(code: str, year: int) -> str:
    return f"day_ahead_price_nbnl_{code}_{year}_eur_p_mwh"


NBNL_PRICE_COLUMNS = [nbnl_price_column(code, year) for code in NBNL_SCENARIOS for year in NBNL_TARGET_YEARS]
for _col in NBNL_PRICE_COLUMNS:
    # Absolute EUR/MWh, not normalized - same convention as day_ahead_price_eur_p_mwh,
    # but distinguished from "none" (out-of-scope/NaN) since these ARE populated.
    NORMALIZATION[_col] = "none_absolute"

# Day-ahead/market prices are set per hour, not smoothly varying - §5.5 confirmed
# day_ahead_price_eur_p_mwh is held constant across all 4 quarters of an hour in the
# existing workbook (step function, "interval-start" convention), not interpolated.
# The NBNL price columns get the same treatment for the same reason.
STEP_HOLD_COLUMNS = {"day_ahead_price_eur_p_mwh", *NBNL_PRICE_COLUMNS}

# The full sheet layout for profiles_weather{YEAR}[_h]: the standard COLUMN_ORDER (matching
# profiles_2025's layout, per the original instructions) plus the 8 subsector heat-demand
# columns and the 16 NBNL price columns, appended at the end. This is a deliberate
# deviation from "match the exact column layout" - added on request, kept as an addition
# rather than a replacement so industry_other_h_demand (the aggregate) and
# day_ahead_price_eur_p_mwh (NaN, out of scope) are still present for anything expecting
# the original layout.
WEATHER_SHEET_COLUMNS = (
    COLUMN_ORDER + EXTRA_SUBSECTOR_COLUMNS + EXTRA_HOUSEHOLD_COLUMNS + EXTRA_WIND_COLUMNS + NBNL_PRICE_COLUMNS
)


def _set_pyetm_environment(env: str) -> None:
    """Switch pyetm's target engine and drop cached config/client state.

    reload_configuration() alone is NOT enough - pyetm.get_client() is an
    lru_cache(maxsize=1) that is never cleared by reload_configuration(), so the
    first BaseClient built in a process stays cached (with its base_url baked in)
    for the rest of the process unless explicitly cleared here too. Verified this
    was actually necessary by testing a bare reload_configuration()-only toggle
    first - it silently kept using the first environment's base_url.
    """
    if env:
        os.environ["ENVIRONMENT"] = env
    else:
        os.environ.pop("ENVIRONMENT", None)
    reload_configuration()
    pyetm.get_client.cache_clear()


def _industry_steel_e_cols(electricity_profiles: pd.DataFrame) -> list[str]:
    return [c for c in electricity_profiles.columns if c.startswith("industry_steel_") and c.endswith(".input (MW)")]


def _industry_other_e_cols(electricity_profiles: pd.DataFrame, steel_cols: list[str]) -> list[str]:
    return [
        c
        for c in electricity_profiles.columns
        if c.startswith("industry_") and c.endswith(".input (MW)") and c not in steel_cols
    ]


def _log_subsector_shape_diagnostics(label: str, series_by_col: dict[str, pd.Series]) -> None:
    """Warn when subsector heat-demand columns turn out not to be independent shapes.

    Discovered by inspection (see NOTES.md): under this scenario's default technology
    mix, several industry heat subsectors are flat (constant, no hourly variation -
    plausible for continuous-process industries like steel/aluminium/chemicals), and
    others share one identical normalized shape template scaled by magnitude. Adding
    a per-run check so this doesn't require re-discovering by hand, and so a future
    scenario config that changes this (e.g. different technology shares) gets noticed.
    """
    flat = [col for col, s in series_by_col.items() if s.std() == 0 or (s.mean() and s.std() / s.mean() < 1e-6)]
    if flat:
        log.warning("%s: these industry heat subsector columns are flat/constant (no hourly variation): %s", label, flat)

    names = [c for c in series_by_col if c not in flat]
    groups: list[list[str]] = []
    seen: set[str] = set()
    for i, a in enumerate(names):
        if a in seen:
            continue
        group = [a]
        for b in names[i + 1 :]:
            if b in seen:
                continue
            if series_by_col[a].sum() > 0 and series_by_col[b].sum() > 0:
                corr = np.corrcoef(series_by_col[a], series_by_col[b])[0, 1]
                if corr > 0.9999:
                    group.append(b)
                    seen.add(b)
        if len(group) > 1:
            groups.append(group)
        seen.add(a)
    for group in groups:
        log.warning(
            "%s: these industry heat subsector columns share an identical normalized shape "
            "(same underlying ETM demand-shape template, scaled by magnitude only): %s",
            label,
            group,
        )


def normalize(series: pd.Series, kind: str) -> pd.Series:
    if kind == "max":
        peak = series.max()
        if peak == 0:
            log.warning("series is all-zero, cannot normalize by max: %s", series.name)
            return series
        return series / peak
    if kind == "sum":
        total = series.sum()
        if total == 0:
            log.warning("series is all-zero, cannot normalize by sum: %s", series.name)
            return series
        return series / total
    return series


def fetch_weather_year(label: str) -> dict[str, pd.Series | None]:
    """Fetch and map one weather year's ETM curves to COLUMN_ORDER (minus t_h).

    Returns a dict of column name -> raw (un-normalized) hourly pd.Series, or
    None for columns that could not be sourced from ETM for this pull (logged
    as a warning, not silently skipped).
    """
    log.info("Building scenario for weather year %r (area_code=%s, end_year=%s)", label, AREA_CODE, END_YEAR)
    session = Session.new(area_code=AREA_CODE, end_year=END_YEAR)
    session.update_user_values({WEATHER_INPUT_KEY: label})

    electricity_profiles = session.get_hourly_curve("electricity_profiles")
    household_heat = session.get_hourly_curve("household_heat")
    buildings_heat = session.get_hourly_curve("buildings_heat")
    network_gas_profiles = session.get_hourly_curve("network_gas_profiles")
    hydrogen_profiles = session.get_hourly_curve("hydrogen_profiles")

    cols: dict[str, pd.Series | None] = {}

    cols["ambientTemperature_degC"] = None
    cols["day_ahead_price_eur_p_mwh"] = None
    cols["co2_factor_kg_per_kwh"] = None
    log.warning(
        "%s: ambientTemperature_degC / day_ahead_price_eur_p_mwh / co2_factor_kg_per_kwh "
        "are out of scope for ETM weather years (no historical NL market/CO2 data for "
        "1987/1997/2004, and these aren't ETM-sourced concepts) - left as NaN, per §7.",
        label,
    )

    wind = electricity_profiles[WIND_INLAND_COL]
    cols["wind_e_prod_normalized"] = wind
    cols["wind_e_prod_normalized_hvh"] = wind
    cols["wind_e_prod_normalized_heibloem"] = wind
    log.info(
        "%s: ETM has no per-location wind curves (national aggregate only) - "
        "using the same %s curve for all 3 wind columns.",
        label,
        WIND_INLAND_COL,
    )

    cols["wind_e_prod_normalized_offshore"] = electricity_profiles[WIND_OFFSHORE_COL]

    solar = electricity_profiles[SOLAR_COLS].sum(axis=1)
    solar.name = "solar_total"
    cols["solar_e_prod_south35deg_normalized"] = solar
    cols["solar_e_prod_eastwest15deg_normalized"] = solar
    log.info(
        "%s: ETM has no orientation split for solar - using the same aggregate "
        "curve (sum of %d segment curves) for both orientation columns.",
        label,
        len(SOLAR_COLS),
    )

    cols["house_e_demand_other"] = electricity_profiles[HOUSE_E_OTHER_COLS].sum(axis=1)
    cols["building_e_demand_other"] = electricity_profiles[BUILDING_E_OTHER_COLS].sum(axis=1)
    cols["house_cooking_demand"] = electricity_profiles[HOUSE_COOKING_COL]
    cols["logistics_fleet_e_hgv"] = electricity_profiles[LOGISTICS_HGV_COL]

    steel_cols = _industry_steel_e_cols(electricity_profiles)
    other_cols = _industry_other_e_cols(electricity_profiles, steel_cols)
    log.info("%s: industry_steel_e_demand <- sum of %d columns: %s", label, len(steel_cols), steel_cols)
    log.info("%s: industry_other_e_demand <- sum of %d columns", label, len(other_cols))
    cols["industry_steel_e_demand"] = electricity_profiles[steel_cols].sum(axis=1)
    cols["industry_other_e_demand"] = electricity_profiles[other_cols].sum(axis=1)

    cols["industry_steel_h_demand"] = (
        network_gas_profiles[INDUSTRY_STEEL_H_GAS_COLS].sum(axis=1)
        + hydrogen_profiles[INDUSTRY_STEEL_H_H2_COLS].sum(axis=1)
    )

    other_h_total = None
    subsector_series_by_col: dict[str, pd.Series] = {"industry_steel_h_demand": cols["industry_steel_h_demand"]}
    for subsector_col, gas_cols, h2_cols in INDUSTRY_OTHER_H_SUBSECTORS:
        subsector_series = network_gas_profiles[gas_cols].sum(axis=1) + hydrogen_profiles[h2_cols].sum(axis=1)
        cols[subsector_col] = subsector_series
        subsector_series_by_col[subsector_col] = subsector_series
        other_h_total = subsector_series if other_h_total is None else other_h_total + subsector_series
    cols["industry_other_h_demand"] = other_h_total

    _log_subsector_shape_diagnostics(label, subsector_series_by_col)

    log.info(
        "%s: industry_steel_h_demand <- network_gas+hydrogen final demand for metal_steel "
        "(2 cols); industry_other_h_demand <- sum of %d subsector columns "
        "(%s), excluding _non_energetic (feedstock) and shared CHP nodes.",
        label,
        len(INDUSTRY_OTHER_H_SUBSECTORS),
        ", ".join(EXTRA_SUBSECTOR_COLUMNS),
    )

    if HOUSE_HOT_WATER_COL in household_heat.columns:
        cols["house_h_demand_hot_water"] = household_heat[HOUSE_HOT_WATER_COL]
    else:
        cols["house_h_demand_hot_water"] = None
        log.warning("%s: %s not found in household_heat curve", label, HOUSE_HOT_WATER_COL)

    missing_space_heating = [c for c in HOUSE_SPACE_HEATING_COLS if c not in household_heat.columns]
    discovered_space_heating = [
        c for c in household_heat.columns if "space_heating" in c and c.endswith(".input.demand (MW)")
    ]
    if set(discovered_space_heating) != set(HOUSE_SPACE_HEATING_COLS):
        log.warning(
            "%s: HOUSE_SPACE_HEATING_COLS (%d cols) doesn't match what's actually in "
            "household_heat (%d cols matching the space_heating/.input.demand pattern) - "
            "missing: %s, unexpected extra: %s",
            label,
            len(HOUSE_SPACE_HEATING_COLS),
            len(discovered_space_heating),
            missing_space_heating,
            set(discovered_space_heating) - set(HOUSE_SPACE_HEATING_COLS),
        )
    if missing_space_heating:
        cols["house_h_demand_space_heating"] = None
    else:
        cols["house_h_demand_space_heating"] = household_heat[HOUSE_SPACE_HEATING_COLS].sum(axis=1)
        log.info(
            "%s: house_h_demand_space_heating <- sum of %d household_heat columns "
            "(HOUSE_SPACE_HEATING_COLS: households_useful_demand_for_space_heating_* "
            "across house types/construction eras)",
            label,
            len(HOUSE_SPACE_HEATING_COLS),
        )

    missing_building_heat = [c for c in BUILDING_HEAT_COLS if c not in buildings_heat.columns]
    if missing_building_heat:
        cols["building_h_demand"] = None
        log.warning("%s: building_h_demand columns missing: %s", label, missing_building_heat)
    else:
        cols["building_h_demand"] = buildings_heat[BUILDING_HEAT_COLS].sum(axis=1)

    return cols


def fetch_nbnl_prices_multi(weather_years: list[str]) -> dict[str, dict[str, np.ndarray]]:
    """Fetch ETM's own market-clearing price for each of the 4 NBNL 2025 scenarios x
    4 target years, across the given weather-year settings. Returns
    {weather_year: {column_name: 8760 values}}.

    Session.new(area_code=..., end_year=..., template_id=X) does NOT work here - tested
    it first and it silently created a blank default scenario (0 user_values) instead of
    copying the ~1130 values that actually distinguish these scenarios. Confirmed by
    checking two different NBNL scenarios' resulting electricity_price curves were
    numerically identical when built that way - a real bug caught before it could
    silently produce 16 copies of the same "generic default" price curve. The fix is
    Session.load(template_id).copy_with_preset(), which correctly inherits all
    user_values (verified: 1130 in, 1130 out) into a private, mutable copy - so
    overriding settings_weather_curve_set and modifying it doesn't touch the original
    published scenario.

    Loads and copies each of the 16 (scenario, year) templates exactly ONCE, then
    iterates over all requested weather_years on that same copy - not once per weather
    year (16 loads+copies total, not 16 x len(weather_years)). get_hourly_curve()
    caches by default, so clear_hourly_curves_cache() must be called after each
    update_user_values() or it silently returns the previous weather year's curve
    (caught by testing: all weather years returned an identical price series before
    adding the cache-clear call).

    Also removes NBNL_WEATHER_CUSTOM_CURVES from the copy before setting
    settings_weather_curve_set - without this, every NBNL scenario's own pinned wind/
    solar/temperature/heating curves silently override the weather-year input entirely
    (see NBNL_WEATHER_CUSTOM_CURVES for how this was caught and confirmed).
    """
    _set_pyetm_environment(NBNL_ENGINE_ENVIRONMENT)
    try:
        result: dict[str, dict[str, np.ndarray]] = {wy: {} for wy in weather_years}
        for code, info in NBNL_SCENARIOS.items():
            for year in NBNL_TARGET_YEARS:
                template_id = info["years"][year]
                col = nbnl_price_column(code, year)
                tmpl = Session.load(template_id)
                n_user_values = len(tmpl.user_values())
                if n_user_values < 100:
                    log.warning(
                        "NBNL %s %s (template_id=%s): only %d user_values on the loaded "
                        "template (expected ~1130) - this scenario may not be a real NBNL "
                        "preset; treat %s with suspicion.",
                        code,
                        year,
                        template_id,
                        n_user_values,
                        col,
                    )
                copy = tmpl.copy_with_preset()
                present_weather_curves = {c.key for c in copy.custom_curves.curves} & set(NBNL_WEATHER_CUSTOM_CURVES)
                if present_weather_curves != set(NBNL_WEATHER_CUSTOM_CURVES):
                    log.warning(
                        "NBNL %s %s (template_id=%s): expected weather/* custom curves %s, "
                        "found %s - settings_weather_curve_set may be partially or fully "
                        "overridden still for this scenario.",
                        code,
                        year,
                        template_id,
                        sorted(NBNL_WEATHER_CUSTOM_CURVES),
                        sorted(present_weather_curves),
                    )
                if present_weather_curves:
                    copy.remove_custom_curves(present_weather_curves)

                # copy_with_preset() + remove_custom_curves() has a race condition:
                # a handful of the template's own values (always including
                # flexibility_outdoor_temperature, plus 2 of the 3 flh_of_wind_*/
                # flh_of_solar_pv_solar_radiation FLH scalars - WHICH 2 varies run to
                # run) silently come back missing on the copy instead of being carried
                # over. Caught by getting 3 different stable electricity_price means
                # (105.64, 105.64, 113.35, 121.87...) across otherwise-identical fresh
                # copies of the same template/weather-year - confirmed root cause by
                # diffing copy.user_values() against the template's and finding exactly
                # these keys missing, varying which ones each time. Fix: after setup,
                # diff every value against the template and restore anything missing
                # before proceeding - not just the specific keys observed here, in case
                # this race condition can drop other values too. Verified: 4 repeated
                # fresh copies all converged on the identical mean price after this.
                template_values = tmpl.user_values()
                copy_values = copy.user_values()
                dropped = {
                    k: v for k, v in template_values.items() if k != WEATHER_INPUT_KEY and copy_values.get(k) != v
                }
                if dropped:
                    log.warning(
                        "NBNL %s %s (template_id=%s): copy_with_preset() dropped %d value(s) "
                        "that should have carried over from the template - restoring before "
                        "proceeding: %s",
                        code,
                        year,
                        template_id,
                        len(dropped),
                        sorted(dropped.keys()),
                    )
                    copy.update_user_values(dropped)

                for weather_year in weather_years:
                    copy.update_user_values({WEATHER_INPUT_KEY: weather_year})
                    copy.clear_hourly_curves_cache()
                    df = copy.get_hourly_curve("electricity_price")
                    values = df.iloc[:, 0].to_numpy(dtype=float)
                    if len(values) != 8760:
                        raise ValueError(f"NBNL {code} {year} ({weather_year}): expected 8760 rows, got {len(values)}")
                    result[weather_year][col] = values
                    log.info(
                        "%s: %s <- NBNL '%s' %s (template_id=%s, area=%s), electricity_price, mean=%.2f EUR/MWh",
                        weather_year,
                        col,
                        info["name"],
                        year,
                        template_id,
                        tmpl.area_code,
                        values.mean(),
                    )
        return result
    finally:
        _set_pyetm_environment("")


def build_hourly_df(label: str) -> pd.DataFrame:
    raw = fetch_weather_year(label)
    n_hours = 8760
    data = {"t_h": np.arange(n_hours, dtype=float)}
    for col in WEATHER_SHEET_COLUMNS[1:]:
        series = raw.get(col)
        if series is None:
            data[col] = np.full(n_hours, np.nan)
            continue
        values = series.to_numpy(dtype=float)
        if len(values) != n_hours:
            raise ValueError(f"{label}/{col}: expected {n_hours} hourly rows, got {len(values)}")
        data[col] = normalize(pd.Series(values, name=col), NORMALIZATION[col]).to_numpy()
    return pd.DataFrame(data, columns=WEATHER_SHEET_COLUMNS)


def interpolate_to_quarter_hour(hourly_df: pd.DataFrame) -> pd.DataFrame:
    """Linearly interpolate an hourly df to 15-min steps.

    NOTE: this is an interpolation, not an independent resample - see the
    module docstring and NOTES.md for why that's a known divergence from how
    profiles_2025 relates to profiles_2025_h (ETM exposes hourly curves only).
    The last quarter-hour block (t_h in [8759, 8759.75]) is extrapolated
    flat from the final hourly value since there's no hour 8760 to interpolate
    toward - matches how a real calendar year has no "hour after the last".
    """
    n_hours = len(hourly_df)
    t_h_quarter = np.arange(0, n_hours, 0.25)
    data = {"t_h": t_h_quarter}
    hourly_t = hourly_df["t_h"].to_numpy()
    for col in WEATHER_SHEET_COLUMNS[1:]:
        y = hourly_df[col].to_numpy()
        if np.isnan(y).all():
            data[col] = np.full(len(t_h_quarter), np.nan)
            continue
        if col in STEP_HOLD_COLUMNS:
            data[col] = np.repeat(y, 4)
            continue
        y_extended = np.append(y, y[-1])
        t_extended = np.append(hourly_t, hourly_t[-1] + 1)
        data[col] = np.interp(t_h_quarter, t_extended, y_extended)
    return pd.DataFrame(data, columns=WEATHER_SHEET_COLUMNS)


def sanity_check_default_vs_existing(default_hourly: pd.DataFrame, source_xlsx: str) -> None:
    """§6: compare ETM 'default' weather wind/solar shape against profiles_2023."""
    existing = pd.read_excel(source_xlsx, sheet_name="profiles_2023_h")
    for col, etm_col in [
        ("wind_e_prod_normalized", "wind_e_prod_normalized"),
        ("solar_e_prod_south35deg_normalized", "solar_e_prod_south35deg_normalized"),
    ]:
        n = min(len(existing), len(default_hourly))
        corr = np.corrcoef(existing[col].to_numpy()[:n], default_hourly[etm_col].to_numpy()[:n])[0, 1]
        etm_peak_hour = int(default_hourly[etm_col].idxmax())
        existing_peak_hour = int(existing[col].iloc[:n].idxmax())
        log.info(
            "Sanity check %s: correlation(ETM default, profiles_2023_h)=%.3f, "
            "ETM peak hour=%d, profiles_2023_h peak hour=%d",
            col,
            corr,
            etm_peak_hour,
            existing_peak_hour,
        )


def write_workbook(hourly_by_year: dict[str, pd.DataFrame], quarter_by_year: dict[str, pd.DataFrame]) -> None:
    wb = openpyxl.load_workbook(SOURCE_XLSX)
    original_sheets = set(wb.sheetnames)

    for label in ["1987", "1997", "2004"]:
        wb.create_sheet(f"profiles_weather{label}_h")
        wb.create_sheet(f"profiles_weather{label}")

    for label in ["1987", "1997", "2004"]:
        ws_h = wb[f"profiles_weather{label}_h"]
        for row in _df_to_rows(hourly_by_year[label]):
            ws_h.append(row)
        ws_q = wb[f"profiles_weather{label}"]
        for row in _df_to_rows(quarter_by_year[label]):
            ws_q.append(row)

    doc_ws = wb.create_sheet("Documentation_weather_years")
    doc_ws.append(["Profile", "Definition", "Convention", "File", "Source"])
    retrieval_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    scenario_desc = f"pyetm Session.new(area_code={AREA_CODE!r}, end_year={END_YEAR}); {WEATHER_INPUT_KEY}=<year>"
    subsector_definitions = {
        "industry_chemicals_fertilizers_h_demand": "Heat demand (network gas + hydrogen burner) for the chemicals/fertilizers subsector",
        "industry_chemicals_other_h_demand": "Heat demand (network gas + hydrogen burner) for the other-chemicals subsector",
        "industry_chemicals_refineries_h_demand": "Heat demand (network gas + hydrogen burner) for the refineries subsector",
        "industry_metal_aluminium_h_demand": "Heat demand (network gas + hydrogen final demand) for the aluminium subsector",
        "industry_metal_other_metals_h_demand": "Heat demand (network gas + hydrogen final demand) for the other-metals subsector",
        "industry_other_non_specified_h_demand": "Heat demand (network gas + hydrogen final demand, energetic use only) for the non-specified-industry subsector",
        "industry_food_h_demand": "Heat demand (network gas + hydrogen burner) for the food industry subsector",
        "industry_paper_h_demand": "Heat demand (network gas + hydrogen burner) for the paper industry subsector",
    }
    household_definitions = {
        "house_h_demand_space_heating": (
            "Household space heating demand - sum of households_useful_demand_for_space_heating_* "
            "in household_heat across all house types (apartments/terraced/semi_detached/detached) "
            "and construction eras (future/2005_present/1985_2004/1965_1984/1945_1964/before_1945), "
            "24 columns total. Added as an extra column beyond the standard layout, on request - the "
            "original layout covers household hot water and building space heating but never "
            "household space heating, normally the largest household heat end-use. See NOTES.md."
        ),
        "wind_e_prod_normalized_offshore": (
            f"Offshore wind production, sourced from {WIND_OFFSHORE_COL} - a genuinely separate "
            "ETM production node from wind_e_prod_normalized (inland), unlike _hvh/_heibloem which "
            "are duplicates of the inland curve (ETM has no per-location split for those). Added "
            "as an extra column beyond the standard layout, on request. See NOTES.md."
        ),
    }
    nbnl_definitions = {}
    for code, info in NBNL_SCENARIOS.items():
        for year in NBNL_TARGET_YEARS:
            col = nbnl_price_column(code, year)
            nbnl_definitions[col] = (
                f"ETM market-clearing price (electricity_price curve) for the NBNL 2025 "
                f"'{info['name']}' scenario, {year}. Proxy for future EPEX day-ahead price - "
                "ETM's own scenario-specific market-optimization result, not historical EPEX "
                "data (day_ahead_price_eur_p_mwh only covers 2023-2025). Source: scenario "
                f"template_id={info['years'][year]} on the pinned '2025-01' ETM engine version "
                "(energytransitionmodel.com 'Featured scenarios', published by Netbeheer "
                "Nederland), copied via Session.load(template_id).copy_with_preset() then "
                f"settings_weather_curve_set={{'1987','1997','2004'}} applied."
            )

    for col in WEATHER_SHEET_COLUMNS[1:]:
        norm = NORMALIZATION[col]
        conv = {
            "max": "Normalized so MAX = 1 (capacity-factor style)",
            "sum": "Normalized so annual sum = 1 (load-shape style)",
            "none": "Not sourced from ETM for weather years - left blank (NaN), see NOTES.md",
            "none_absolute": (
                "Absolute value (EUR/MWh), not normalized. Quarter-hour sheet holds each "
                "hour's value constant across its 4 quarters (step function), not linearly "
                "interpolated - matches day_ahead_price_eur_p_mwh's own convention."
            ),
        }[norm]
        if col in subsector_definitions:
            definition = (
                f"{subsector_definitions[col]}. One of 8 subsectors that together sum to "
                "industry_other_h_demand (added as extra columns beyond the standard layout, "
                "on request - see NOTES.md)."
            )
        elif col in nbnl_definitions:
            definition = nbnl_definitions[col]
        elif col in household_definitions:
            definition = household_definitions[col]
        else:
            definition = f"Weather-year variant of {col} for profiles_weather{{1987,1997,2004}}[_h]"
        doc_ws.append(
            [
                col,
                definition,
                conv,
                "db_profiles_loader/fetch_etm_weather_years.py",
                scenario_desc + f" (retrieved {retrieval_date})",
            ]
        )

    for name in wb.sheetnames:
        if name not in original_sheets and name != "Documentation_weather_years":
            continue

    wb.save(OUTPUT_XLSX)
    log.info("Wrote %s (original sheets preserved: %s)", OUTPUT_XLSX, sorted(original_sheets))


def _df_to_rows(df: pd.DataFrame):
    yield list(df.columns)
    for row in df.itertuples(index=False):
        yield list(row)


def main() -> None:
    non_default_years = [wy for wy in WEATHER_YEARS if wy != "default"]
    nbnl_prices_by_year = fetch_nbnl_prices_multi(non_default_years)

    hourly_by_year = {}
    quarter_by_year = {}
    for label in WEATHER_YEARS:
        hourly = build_hourly_df(label)
        if label != "default":
            for col, values in nbnl_prices_by_year[label].items():
                hourly[col] = values
            hourly_by_year[label] = hourly
            quarter_by_year[label] = interpolate_to_quarter_hour(hourly)
        else:
            sanity_check_default_vs_existing(hourly, SOURCE_XLSX)

    write_workbook(hourly_by_year, quarter_by_year)


if __name__ == "__main__":
    main()
